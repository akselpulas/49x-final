"""
CE49X Final Project - RSS Feed Article Collector
Enhanced version with more feeds and Excel output
"""

import feedparser
import pandas as pd
from pathlib import Path
from datetime import datetime
import re

# Fix Unicode encoding for Windows console
import sys
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

# -----------------------------
# 1. RSS FEED KAYNAKLARI (Genişletilmiş)
# -----------------------------
RSS_FEEDS = {
    # Industry News Portals
    "ENR": "https://www.enr.com/rss",
    "ConstructionDive": "https://www.constructiondive.com/rss",
    "ConstructionDiveTech": "https://www.constructiondive.com/topic/technology/feed/",
    "CivilEngineer": "https://www.csemag.com/rss",
    "CivilEngineerMag": "https://csengineermag.com/feed/",
    "NewCivilEngineer": "https://www.newcivilengineer.com/feed/",
    "ConstructionIndex": "https://www.theconstructionindex.co.uk/feed/",
    "AECMagazine": "https://aecmag.com/feed/",
    "BIMplus": "https://www.bimplus.co.uk/feed/",
    "BIM42": "https://bim42.com/feed/",
    "CEMEXVentures": "https://cemexventures.com/blog/feed/",
    "EngineeringNews": "https://www.engineeringnews.co.za/rss",
    
    # Tech News (Filtered for construction/infrastructure keywords)
    "TechCrunch": "https://techcrunch.com/feed/",
    "TechCrunchAI": "https://techcrunch.com/tag/artificial-intelligence/feed/",
    "Wired": "https://www.wired.com/feed/rss",
    "WiredAI": "https://www.wired.com/tag/artificial-intelligence/feed/",
    "VentureBeat": "https://venturebeat.com/feed/",
    "VentureBeatAI": "https://venturebeat.com/ai/feed/",
    "GuardianTech": "https://www.theguardian.com/uk/technology/rss",
    "GuardianAI": "https://www.theguardian.com/technology/artificialintelligenceai/rss",
    "TheVerge": "https://www.theverge.com/rss/index.xml",
    "ArsTechnica": "https://feeds.arstechnica.com/arstechnica/index",
    "MITTechReview": "https://www.technologyreview.com/feed/",
    "IEEE": "https://spectrum.ieee.org/rss",
}

# -----------------------------
# 2. KEYWORD LİSTESİ
# (Civil Engineering + AI)
# -----------------------------
CE_KEYWORDS = [
    "construction", "infrastructure", "bridge", "tunnel",
    "structural", "geotechnical", "transportation", "concrete",
    "civil engineering", "building", "architecture", "engineering",
    "construction site", "construction project", "infrastructure development",
    "structural engineering", "building design", "construction industry",
    "public works", "urban planning", "construction management"
]

AI_KEYWORDS = [
    "artificial intelligence", "machine learning", "computer vision",
    "automation", "robotics", "generative ai", "predictive",
    "neural networks", "deep learning", "AI", "ML",
    "intelligent systems", "automated", "smart systems",
    "machine intelligence", "cognitive computing", "AI-powered"
]

ALL_KEYWORDS = CE_KEYWORDS + AI_KEYWORDS

# -----------------------------
# 3. HELPER FUNCTIONS
# -----------------------------

def find_keywords_in_text(text: str, keywords: list) -> list:
    """Find which keywords appear in text."""
    if not text:
        return []
    text_lower = text.lower()
    found = [kw for kw in keywords if kw.lower() in text_lower]
    return found

def clean_html(text: str) -> str:
    """Remove HTML tags from text."""
    if not text:
        return ""
    # Simple HTML tag removal
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&quot;', '"', text)
    return text.strip()

def parse_date(date_str: str) -> str:
    """Parse and format date."""
    if not date_str:
        return ""
    try:
        # Try to parse common date formats
        from dateutil import parser
        dt = parser.parse(date_str)
        return dt.strftime("%Y-%m-%d")
    except:
        # Return first 10 characters if parsing fails
        return date_str[:10] if len(date_str) >= 10 else date_str

# -----------------------------
# 4. MAKALE TOPLAMA
# -----------------------------
print("=" * 70)
print("RSS Feed'lerden Makale Çekiliyor...")
print("=" * 70)
print()

articles = []
total_entries = 0

for source, url in RSS_FEEDS.items():
    print(f"📡 {source}")
    print(f"   URL: {url}")
    try:
        feed = feedparser.parse(url)
        entries = feed.entries
        
        if not entries:
            print(f"   ⚠️  Makale bulunamadı")
            print()
            continue
        
        count = 0
        for entry in entries:
            title = clean_html(entry.get("title", ""))
            summary = clean_html(entry.get("summary", "") or entry.get("description", ""))
            link = entry.get("link", "")
            published = entry.get("published", "") or entry.get("updated", "")
            
            if not title or not link:
                continue
            
            # Find keywords
            combined_text = f"{title} {summary}".lower()
            ce_found = find_keywords_in_text(combined_text, CE_KEYWORDS)
            ai_found = find_keywords_in_text(combined_text, AI_KEYWORDS)
            
            # Must have at least one CE and one AI keyword (strict filtering)
            if len(ce_found) == 0 or len(ai_found) == 0:
                continue
            
            # Also check if summary is too short (might not be relevant)
            if len(summary) < 100 and len(ce_found) == 0 and len(ai_found) == 0:
                continue
            
            # Parse date
            date_formatted = parse_date(published)
            
            articles.append({
                "Title": title,
                "Published Date": date_formatted,
                "Source": source,
                "URL": link,
                "Summary": summary[:500] if summary else "",  # Limit summary length
                "CE Keywords Found": ", ".join(ce_found),
                "AI Keywords Found": ", ".join(ai_found),
                "Total CE Keywords": len(ce_found),
                "Total AI Keywords": len(ai_found),
            })
            count += 1
            total_entries += 1
        
        print(f"   ✅ {count} relevant makale bulundu (toplam {len(entries)} entry)")
        print()
        
    except Exception as e:
        print(f"   ❌ Hata: {str(e)}")
        print()

# -----------------------------
# 5. DATAFRAME OLUŞTURMA
# -----------------------------
if not articles:
    print("❌ Hiç makale bulunamadı!")
    sys.exit(1)

df = pd.DataFrame(articles)

# Sort by date (newest first)
df['Published Date'] = pd.to_datetime(df['Published Date'], errors='coerce')
df = df.sort_values('Published Date', ascending=False, na_position='last')
df['Published Date'] = df['Published Date'].dt.strftime('%Y-%m-%d').fillna('')

# -----------------------------
# 6. EXCEL KAYIT (Güzel formatlanmış)
# -----------------------------
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = OUTPUT_DIR / f"rss_articles_{timestamp}.xlsx"

# Excel writer with formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Articles', index=False)
    
    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Articles']
    
    # Auto-adjust column widths
    for idx, col in enumerate(df.columns, 1):
        max_length = max(
            df[col].astype(str).map(len).max(),
            len(col)
        )
        # Set column width (with some padding)
        worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 50)
    
    # Freeze first row
    worksheet.freeze_panes = 'A2'
    
    # Make header row bold
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Wrap text for long columns
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.column in [1, 4, 5]:  # Title, URL, Summary columns
                cell.alignment = wrap_alignment

print("=" * 70)
print("✅ İşlem Tamamlandı!")
print("=" * 70)
print(f"📊 Toplam çekilen entry: {total_entries}")
print(f"📝 Keyword ile eşleşen makale: {len(df)}")
print(f"📁 Çıktı dosyası: {output_file.name}")
print(f"📂 Tam yol: {output_file.absolute()}")
print()
print("📈 Kaynaklara göre dağılım:")
source_counts = df['Source'].value_counts()
for source, count in source_counts.items():
    print(f"   {source}: {count} makale")
